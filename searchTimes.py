import sys
from imp import reload
reload(sys)

from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import vocabularyProcessing
import time
import random

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter


browser = webdriver.Chrome()

browser.get('https://www.google.com.tw/')

disorderName = vocabularyProcessing.getMentalDisorderName()
searchList = {}

for name in disorderName:
    searchBar = browser.find_element_by_xpath("//input[@type='text']")
    searchBar.clear()
    searchBar.send_keys(name)
    searchBar.send_keys(Keys.ENTER)
    result = browser.find_element_by_xpath("//div[@id='resultStats']")
    strResult = result.text
    numberOfResult = strResult.split()[1]
    
    searchList[name] = numberOfResult
    time.sleep(random.randrange(1,10))
    
browser.close()

def writeExcel(dataList, fileName):
    
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = '病症'
    ws['B1'] = '搜尋結果'
    ws['C1'] = '取log 以10為基底'
    ws['D1'] = '以6.05為閥值'
    ws['E1'] = '是否為常見疾病'

    i = 1
    for k,v in dataList.items():
        col1 = 'A' + str(i+1)
        col2 = 'B' + str(i+1)
        col3 = 'C' + str(i+1) 
        col4 = 'D' + str(i+1)
        col5 = 'E' + str(i+1)

        ws[col1] = k
        ws[col2] = v
        ws[col3] = '=LOG10(' + col2 + ')'
        ws[col4] = '=IF(' + col3 + '<6.05,0,1)'
        ws[col5] = '=IF(' + col4 + '=1,"常見","不常見")'
        
        i += 1
        
    wb.save(fileName)

writeExcel(searchList, 'mental_disorder.xlsx')