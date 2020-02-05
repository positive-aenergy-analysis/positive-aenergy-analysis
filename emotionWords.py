#統計字數寫入excel
import sys
from imp import reload
reload(sys)
import json
import jieba
import jieba.analyse
#write exl
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
# from openpyxl.cell import get

def readHappyWord():
    
    happyWord = []
    
    with open('vocabulary/happy.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            happyWord.append(data)
    
    return happyWord

def readUnhappyWord():
    
    unhappyWord = []
    
    with open('vocabulary/unhappy.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            unhappyWord.append(data)
    
    return unhappyWord

def check(filepath, wordList):
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    datawords = table.col_values(0)
    datatimes = table.col_values(1)

    remainderWords = []
    remainderWords = list(filter(lambda a: a in wordList and a != '\n', datawords))
    
    timesList = []
    score = 0

    for key in remainderWords:
        for i in range(len(datawords)):
            if key == datawords[i]:
                timesList.append(datatimes[i])

    print(remainderWords,timesList)
    print('------------------------------------------------------------')

    return remainderWords, timesList

def writeTotalExcel(fileName, nameList, happyTimesList, unhappyTimesList):
    
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '檔名'
    ws['B1'] = '正面詞彙總數'
    ws['C1'] = '負面詞彙總數'
    ws['D1'] = '文章正面指數'
    
    for i in range(len(nameList)):
        col1 = 'A' + str(i+2)
        col2 = 'B' + str(i+2)
        col3 = 'C' + str(i+2)
        col4 = 'D' + str(i+2)

        ws[col1] = nameList[i]
        ws[col2] = happyTimesList[i]
        ws[col3] = unhappyTimesList[i]

        if (happyTimesList[i] + unhappyTimesList[i] == 0):
            ws[col4] = 0
        else:
            ws[col4] = happyTimesList[i] / (happyTimesList[i] + unhappyTimesList[i])
        
    wb.save(fileName)

def writeExcel(fileName, happyKeyList, happyTimesList, unhappyKeyList, unhappyTimesList):
    
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '正面詞彙'
    ws['B1'] = '正面詞彙出現次數'
    ws['D1'] = '負面詞彙'
    ws['E1'] = '負面詞彙出現次數'
    
    for i in range(len(happyKeyList)):
        col1 = 'A' + str(i+2)
        col2 = 'B' + str(i+2)

        ws[col1] = happyKeyList[i]
        ws[col2] = happyTimesList[i]

    for i in range(len(unhappyKeyList)):
        col1 = 'D' + str(i+2)
        col2 = 'E' + str(i+2)

        ws[col1] = unhappyKeyList[i]
        ws[col2] = unhappyTimesList[i]
        
    wb.save(fileName)

if __name__ == "__main__":

    json_data = open('data_set/correspondence_table.json', encoding='utf8').read()
    data = json.loads(json_data)

    nameList = []
    happyKeywordTimesList = []
    unhappyKeywordTimesList = []

    for i in data:
        fileName = i['file_name']
        filePath = './data_set_wordCount/wordCount_' + fileName + '.xlsx'
        excelPath = './data_set_emotion/emotion_' + fileName + '.xlsx'
        
        happyKeyList, happyTimesList = check(filePath,readHappyWord())
        unhappyKeyList, unhappyTimesList = check(filePath,readUnhappyWord())

        writeExcel(excelPath, happyKeyList, happyTimesList, unhappyKeyList, unhappyTimesList)

        nameList.append(fileName)
        happyKeywordTimesList.append(sum(happyTimesList))
        unhappyKeywordTimesList.append(sum(unhappyTimesList))


    writeTotalExcel('./data_set_emotion/totalEmotion.xlsx',nameList, happyKeywordTimesList, unhappyKeywordTimesList)
