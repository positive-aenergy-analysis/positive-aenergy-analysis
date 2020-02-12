#統計字數寫入excel
import sys
from imp import reload
reload(sys)
import json
import jieba
import jieba.analyse
#write exl
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
# from openpyxl.cell import get

professionalTerms = {}
fileNameList = []
professionalTermsAppearTimesList = []
totalTermsList = []
professionalTermsProportionList = []
professionalWordCountList = []
totalWordCountList = []
wordCountProportionList = []
diversityScoreList = []

def readProfessionalTermsFile():
    data = xlrd.open_workbook('vocabulary/mental_disorder.xlsx')
    table = data.sheets()[0]
    disorderWords = table.col_values(0)
    disorderScores = table.col_values(2)
    
    return disorderWords, disorderScores

def combineKeyValueOfProfessionalTerms(disorderWords, disorderScores):
    for i in range(1, len(disorderWords)):
        professionalTerms[disorderWords[i]] = disorderScores[i]

def check(filepath, disorderWords, disorderScores):
    # data 為分詞完的文章, dataKeys 是column(A), dataTimes 是column(B)
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    dataKeys = table.col_values(0)
    dataTimes = table.col_values(1)
    dataTotalWords = table.col_values(2)[0]
    dataDictionary = {}
    for i in range(len(dataKeys)):
        dataDictionary[dataKeys[i]] = dataTimes[i]

    professionalTermsOfArticleList = []
    professionalTermsOfArticleList = list(filter(lambda a: a in disorderWords and a != '\n' and a != '病症', dataKeys))
    
    scoreList = []
    appearTimesList = []
    professionalWordCountOfArticleList = []
    score = 0

    # key 對照 appearTimes, 專業詞分數, 專業詞出現次數
    for key in professionalTermsOfArticleList:
        appearTimes = dataDictionary[key]
        scoreList.append(professionalTerms[key])
        appearTimesList.append(appearTimes)
        professionalWordCountOfArticleList.append(len(key) * appearTimes)

    professionalTermsAppearTimesList.append(sum(appearTimesList))
    totalTermsList.append(sum(dataTimes))
    professionalTermsProportionList.append(sum(appearTimesList) / sum(dataTimes))
    professionalWordCountList.append(sum(professionalWordCountOfArticleList))
    totalWordCountList.append(dataTotalWords)
    wordCountProportionList.append(sum(professionalWordCountOfArticleList) / dataTotalWords)

    for i in range(len(professionalTermsOfArticleList)):
        totalAppearTimes = sum(appearTimesList)
        if totalAppearTimes == 0:
            break
        score += pow((appearTimesList[i] / totalAppearTimes), 2)

    diversityScoreList.append(score)

    return professionalTermsOfArticleList, scoreList, appearTimesList, professionalWordCountOfArticleList

def writeExcel(fileName, professionalTermsOfArticleList, scoreList, appearTimesList, professionalWordCountOfArticleList, col_title1, col_title2, col_title3, col_title4):
    
    wb = Workbook()
    ws = wb.active

    ws['A1'] = col_title1
    ws['B1'] = col_title2
    ws['C1'] = col_title3
    ws['D1'] = col_title4
    
    for i in range(len(professionalTermsOfArticleList)):
        col1 = 'A' + str(i+2)
        col2 = 'B' + str(i+2)
        col3 = 'C' + str(i+2)
        col4 = 'D' + str(i+2)

        ws[col1] = professionalTermsOfArticleList[i]
        ws[col2] = scoreList[i]
        ws[col3] = appearTimesList[i]
        ws[col4] = professionalWordCountOfArticleList[i]
        
    wb.save(fileName)

def writeTotalExcel(fileName):
    
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '檔名'
    ws['B1'] = '專業詞總出現次數'
    ws['C1'] = '總詞數'
    ws['D1'] = '專業詞比例'
    ws['E1'] = '專業詞總字數'
    ws['F1'] = '全文總字數'
    ws['G1'] = '字數比例'
    ws['H1'] = '專業詞多樣性分數'
    
    for i in range(len(fileNameList)):
        col1 = 'A' + str(i+2)
        col2 = 'B' + str(i+2)
        col3 = 'C' + str(i+2)
        col4 = 'D' + str(i+2)
        col5 = 'E' + str(i+2)
        col6 = 'F' + str(i+2)
        col7 = 'G' + str(i+2)
        col8 = 'H' + str(i+2)

        ws[col1] = fileNameList[i]
        ws[col2] = professionalTermsAppearTimesList[i]
        ws[col3] = totalTermsList[i]
        ws[col4] = professionalTermsProportionList[i]
        ws[col5] = professionalWordCountList[i]
        ws[col6] = totalWordCountList[i]
        ws[col7] = wordCountProportionList[i]
        ws[col8] = diversityScoreList[i]
        
    wb.save(fileName)

if __name__ == "__main__":

    disorderWords, disorderScores = readProfessionalTermsFile()
    combineKeyValueOfProfessionalTerms(disorderWords, disorderScores)

    json_data = open('data_set/correspondence_table.json', encoding='utf8').read()
    data = json.loads(json_data)

    for i in data:
        fileName = i['file_name']
        filePath = './data_set_wordCount/wordCount_' + fileName + '.xlsx'
        excelPath = './data_set_score/test_score_' + fileName + '.xlsx'

        fileNameList.append(fileName)
        professionalTermsOfArticleList, scoreList, appearTimesList, professionalWordCountOfArticleList = check(filePath,disorderWords, disorderScores)
        writeExcel(excelPath, professionalTermsOfArticleList, scoreList, appearTimesList, professionalWordCountOfArticleList, '精神疾病相關詞彙', '分數', '出現次數', '字數')

    writeTotalExcel('./data_set_score/testtotalScore.xlsx')