#統計字數寫入excel
import sys
from imp import reload
reload(sys)
import json
import jieba
import jieba.analyse
#write exl
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
# from openpyxl.cell import get

def addNewWord():
    with open('vocabulary/mental_disorder.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            jieba.add_word(data)

    with open('vocabulary/happy.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            jieba.add_word(data)

    with open('vocabulary/unhappy.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            jieba.add_word(data)

    with open('vocabulary/stop.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            jieba.add_word(data)

def readStopWord():
    
    stopWords = []
    
    with open('vocabulary/stop.txt', 'r', encoding='utf-8') as file:
        for data in file.readlines():
            data = data.strip()
            stopWords.append(data)
    
    return stopWords

def cutWord(fileName):
    
    wordList = []
    tags = []
    remainderWords = []
    stopWords = readStopWord()
    totalWords = 0

    for line in open(fileName, encoding = 'utf-8'):
        item = line.strip('\n\r').split('，')
        totalWords += len(line)
        for segment in item:
            # jieba 分詞
            tags = jieba.analyse.extract_tags(segment,100000)
            # print(tags)
            remainderWords = list(filter(lambda a: a not in stopWords and a != '\n', tags))
            # print(remainderWords)
            for t in remainderWords:
                wordList.append(t)

    return wordList, totalWords

def wordCount(wordList):

    wordDict = {}
    keyList = []

    with open('wordCount.txt', 'w', encoding='utf-8') as wf2:
        # 開檔
        for item in wordList:
            if item not in wordDict:
                wordDict[item] = 1
            else:
                wordDict[item] += 1

        orderList = list(wordDict.values())
        orderList.sort(reverse=True)
        # print orderList
        for i in range(len(orderList)):
            for key in wordDict:
                if wordDict[key] == orderList[i]:
                    wf2.write(key + ' ' + str(wordDict[key]) + '\n')
                    keyList.append(key)
                    wordDict[key] = 0

    return keyList,orderList

def writeExcel(fileName,keyList, valueList, totalWords):
    
    wb = Workbook()
    ws = wb.active
    
    for i in range(len(keyList)):
        text1 = 'A' + str(i+1)
        text2 = 'B' + str(i+1)

        ws[text1] = keyList[i]
        ws[text2] = orderList[i]
    
    ws['C1'] = totalWords
        
    wb.save(fileName)

if __name__ == "__main__":

    json_data = open('data_set/correspondence_table.json', encoding='utf8').read()
    data = json.loads(json_data)
    
    addNewWord()

    for i in data:
        fileName = i['file_name']
        filePath = './data_set/' + fileName + '.txt'
        excelName = './data_set_wordCount/wordCount_' + fileName + '.xlsx'

        wordList, totalWords = cutWord(filePath)
        keyList,orderList = wordCount(wordList)

        writeExcel(excelName, keyList, orderList, totalWords) 
